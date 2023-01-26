import csv
import json
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from datetime import date
from datetime import timedelta
import Report


def main():
    CUSTOMERS = Report.get_filename("Customer")
    AGED_RECEIVABLES_SUMMARY = Report.get_filename("Aged_Receivables_Summary")
    customerJsonPath = r'customers_data.json'
    report_file_name = "PlanwayIndividualAgeingReport.xlsx"

    # create a workbook to store aging report
    wb = Workbook()
    sheet = wb.active
    output_file_row_num = 1
    output_file_col_num = 1
    col_num_in_alphabets = 64
    string_size = 11
    num_size = 10

    # write title
    # font style = "Arial", size = 20, bold

    sheet.column_dimensions[chr(
        col_num_in_alphabets+output_file_col_num)].width = float(4)
    sheet.cell(row=output_file_row_num, column=1).font = Font(
        name="Arial", size=18, bold=True)
    sheet.cell(row=output_file_row_num,
               column=1).value = "Aged Receivables Summary [INDIVIDUAL]"
    output_file_row_num += 1
    # font style = "Arial", size = 14, no bold
    sheet.cell(row=output_file_row_num, column=1).font = Font(
        name="Arial", size=12, bold=False)
    sheet.cell(row=output_file_row_num,
               column=1).value = "PLANWAY POULTRY INC."
    output_file_row_num += 1
    sheet.cell(row=output_file_row_num, column=1).font = Font(
        name="Arial", size=12, bold=False)
    # Planway aging report is the most recent wednesday
    sheet.cell(row=output_file_row_num,
               column=1).value = "Effective as at EOD {}".format(Report.get_most_recent_weekdays(3))
    output_file_row_num += 1

    sheet.cell(row=output_file_row_num, column=1).font = Font(
        name="Arial", size=12, bold=False)
    sheet.cell(row=output_file_row_num, column=1).value = "Ageing by due date"
    sheet.row_dimensions[output_file_row_num+1].height = float(30)
    output_file_row_num += 2
    # sheet.row_dimensions[output_file_row_num].height = float(30)

    # write a filter row
    sheet.cell(row=output_file_row_num, column=output_file_col_num).border = Border(
        top=Side(style="thin"), bottom=Side(style='thin'))
    sheet.cell(row=output_file_row_num, column=output_file_col_num).alignment = Alignment(
        horizontal='left', vertical='top')
    output_file_col_num += 1
    # Name
    sheet.cell(row=output_file_row_num, column=output_file_col_num).font = Font(
        name="Arial", size=string_size, bold=True)
    sheet.cell(row=output_file_row_num,
               column=output_file_col_num).value = "Name"
    sheet.cell(row=output_file_row_num, column=output_file_col_num).alignment = Alignment(
        horizontal='left', vertical='top')
    sheet.cell(row=output_file_row_num, column=output_file_col_num).border = Border(
        top=Side(style="thin"), bottom=Side(style='thin'))
    sheet.column_dimensions[chr(
        col_num_in_alphabets+output_file_col_num)].width = float(29.91)
    output_file_col_num += 1
    # Code
    sheet.cell(row=output_file_row_num, column=output_file_col_num).font = Font(
        name="Arial", size=string_size, bold=True)
    sheet.cell(row=output_file_row_num,
               column=output_file_col_num).value = "Code"
    sheet.cell(row=output_file_row_num, column=output_file_col_num).alignment = Alignment(
        horizontal='left', vertical='top')
    sheet.cell(row=output_file_row_num, column=output_file_col_num).border = Border(
        top=Side(style="thin"), bottom=Side(style='thin'))
    sheet.column_dimensions[chr(
        col_num_in_alphabets+output_file_col_num)].width = float(7.73)
    output_file_col_num += 1
    # Term
    sheet.cell(row=output_file_row_num, column=output_file_col_num).font = Font(
        name="Arial", size=string_size, bold=True)
    sheet.cell(row=output_file_row_num,
               column=output_file_col_num).value = "Term"
    sheet.cell(row=output_file_row_num, column=output_file_col_num).alignment = Alignment(
        horizontal='left', vertical='top')
    sheet.cell(row=output_file_row_num, column=output_file_col_num).border = Border(
        top=Side(style="thin"), bottom=Side(style='thin'))
    sheet.column_dimensions[chr(
        col_num_in_alphabets+output_file_col_num)].width = float(7.73)
    output_file_col_num += 1
    # Limit
    sheet.cell(row=output_file_row_num, column=output_file_col_num).font = Font(
        name="Arial", size=string_size, bold=True)
    sheet.cell(row=output_file_row_num,
               column=output_file_col_num).value = "Limit"
    sheet.cell(row=output_file_row_num, column=output_file_col_num).alignment = Alignment(
        horizontal='right', vertical='top')
    sheet.cell(row=output_file_row_num, column=output_file_col_num).border = Border(
        top=Side(style="thin"), bottom=Side(style='thin'))
    sheet.column_dimensions[chr(
        col_num_in_alphabets+output_file_col_num)].width = float(12.91)
    output_file_col_num += 1
    # Current
    sheet.cell(row=output_file_row_num, column=output_file_col_num).font = Font(
        name="Arial", size=num_size, bold=True)
    sheet.cell(row=output_file_row_num,
               column=output_file_col_num).value = "Current"
    sheet.cell(row=output_file_row_num, column=output_file_col_num).alignment = Alignment(
        horizontal='right', vertical='top')
    sheet.cell(row=output_file_row_num, column=output_file_col_num).border = Border(
        top=Side(style="thin"), bottom=Side(style='thin'))
    sheet.column_dimensions[chr(
        col_num_in_alphabets+output_file_col_num)].width = float(12.91)
    num_of_balance_weeks = 5
    # Week Overdue
    for i in range(1, num_of_balance_weeks):

        sheet.cell(row=output_file_row_num, column=output_file_col_num +
                   i).font = Font(name="Arial", size=num_size, bold=True)

        sheet.cell(row=output_file_row_num, column=output_file_col_num +
                   i).border = Border(top=Side(style="thin"), bottom=Side(style='thin'))
        sheet.cell(row=output_file_row_num, column=output_file_col_num +
                   i).alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
        sheet.cell(row=output_file_row_num, column=output_file_col_num +
                   i).value = f"{i} Week\nOverdue" if i == 1 else f"{i} Weeks\nOverdue"
        sheet.column_dimensions[chr(
            col_num_in_alphabets+output_file_col_num+i)].width = float(12.91)
        i += 1

    output_file_col_num += num_of_balance_weeks
    # Older
    sheet.cell(row=output_file_row_num, column=output_file_col_num).font = Font(
        name="Arial", size=num_size, bold=True)
    sheet.cell(row=output_file_row_num,
               column=output_file_col_num).value = "Older"
    sheet.cell(row=output_file_row_num, column=output_file_col_num).alignment = Alignment(
        horizontal='right', vertical='top')
    sheet.cell(row=output_file_row_num, column=output_file_col_num).border = Border(
        top=Side(style="thin"), bottom=Side(style='thin'))
    sheet.column_dimensions[chr(
        col_num_in_alphabets+output_file_col_num)].width = float(12.91)
    output_file_col_num += 1
    # Total
    sheet.cell(row=output_file_row_num, column=output_file_col_num).font = Font(
        name="Arial", size=num_size, bold=True)
    sheet.cell(row=output_file_row_num,
               column=output_file_col_num).value = "Total"
    sheet.cell(row=output_file_row_num, column=output_file_col_num).alignment = Alignment(
        horizontal='right', vertical='top')
    sheet.cell(row=output_file_row_num, column=output_file_col_num).border = Border(
        top=Side(style="thin"), bottom=Side(style='thin'))
    sheet.column_dimensions[chr(
        col_num_in_alphabets+output_file_col_num)].width = float(12.91)

    # write aging report content
    Report.generate_aging_report(report_file_name, wb, sheet, CUSTOMERS,
                                 customerJsonPath, AGED_RECEIVABLES_SUMMARY, output_file_row_num, "Arial", 11, 30)
    os.remove(customerJsonPath)


if __name__ == "__main__":
    main()
