import csv
import json
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
import datetime
from datetime import date
from datetime import timedelta
import Report


def main():
    CUSTOMERS = Report.get_filename("Customer")
    AGED_RECEIVABLES_SUMMARY = Report.get_filename("Aged_Receivables_Summary")
    customerJsonPath = r'tsw_customers_data.json'
    report_file_name = "TSWMonthlyAgeingReport.xlsx"

    # create a workbook to store aging report
    wb = Workbook()
    sheet = wb.active
    output_file_row_num = 1
    output_file_col_num = 1
    col_num_in_alphabets = 64
    string_size = 11
    num_size = 11
    text_style = "Arial"
    cell_border = "hair"
    today = date.today()

    # write title
    # font style = "Arial", size = 20, bold

    sheet.column_dimensions[chr(
        col_num_in_alphabets+output_file_col_num)].width = float(15)
    sheet.cell(row=output_file_row_num, column=1).font = Font(
        name=text_style, size=20, bold=True)
    sheet.cell(row=output_file_row_num,
               column=1).value = "Aged Receivables Summary"
    output_file_row_num += 1
    # font style = "Arial", size = 14, no bold
    sheet.cell(row=output_file_row_num, column=1).font = Font(
        name=text_style, size=14, bold=False)
    sheet.cell(row=output_file_row_num,
               column=1).value = "TORONTO SUN WAH TRADING LTD."
    output_file_row_num += 1
    sheet.cell(row=output_file_row_num, column=1).font = Font(
        name=text_style, size=14, bold=False)
    # Planway aging report is the most recent wednesday
    sheet.cell(row=output_file_row_num,
               column=1).value = f"As at {today.strftime('%d')} {today.strftime('%B')} {today.strftime('%Y')}"
    output_file_row_num += 1

    sheet.cell(row=output_file_row_num, column=1).font = Font(
        name=text_style, size=14, bold=False)
    sheet.cell(row=output_file_row_num,
               column=1).value = "Ageing by invoice date"
    sheet.row_dimensions[output_file_row_num+1].height = float(30)
    output_file_row_num += 2
    # sheet.row_dimensions[output_file_row_num].height = float(30)

    # write a filter row
    # sheet.cell(row=output_file_row_num, column=output_file_col_num).border = Border(
    #     top=Side(style=cell_border), bottom=Side(style=cell_border))
    # sheet.cell(row=output_file_row_num, column=output_file_col_num).alignment = Alignment(
    #     horizontal='left', vertical='top')
    # output_file_col_num += 1

    sheet.cell(row=output_file_row_num, column=output_file_col_num).font = Font(
        name=text_style, size=string_size, bold=True)
    sheet.cell(row=output_file_row_num,
               column=output_file_col_num).value = "Type"
    sheet.cell(row=output_file_row_num, column=output_file_col_num).border = Border(
        top=Side(style=cell_border), bottom=Side(style=cell_border))
    sheet.cell(row=output_file_row_num, column=output_file_col_num).alignment = Alignment(
        horizontal='left', vertical='top')
    sheet.row_dimensions[output_file_row_num].height = float(30)        
    output_file_col_num += 1

    # Name
    sheet.cell(row=output_file_row_num, column=output_file_col_num).font = Font(
        name=text_style, size=string_size, bold=True)
    sheet.cell(row=output_file_row_num,
               column=output_file_col_num).value = "Name"
    sheet.cell(row=output_file_row_num, column=output_file_col_num).alignment = Alignment(
        horizontal='left', vertical='top')
    sheet.cell(row=output_file_row_num, column=output_file_col_num).border = Border(
        top=Side(style=cell_border), bottom=Side(style=cell_border))
    sheet.column_dimensions[chr(
        col_num_in_alphabets+output_file_col_num)].width = float(29.91)
    output_file_col_num += 1
    # Code
    sheet.cell(row=output_file_row_num, column=output_file_col_num).font = Font(
        name=text_style, size=string_size, bold=True)
    sheet.cell(row=output_file_row_num,
               column=output_file_col_num).value = "Code"
    sheet.cell(row=output_file_row_num, column=output_file_col_num).alignment = Alignment(
        horizontal='left', vertical='top')
    sheet.cell(row=output_file_row_num, column=output_file_col_num).border = Border(
        top=Side(style=cell_border), bottom=Side(style=cell_border))
    sheet.column_dimensions[chr(
        col_num_in_alphabets+output_file_col_num)].width = float(7.73)
    output_file_col_num += 1
    # Term
    sheet.cell(row=output_file_row_num, column=output_file_col_num).font = Font(
        name=text_style, size=string_size, bold=True)
    sheet.cell(row=output_file_row_num,
               column=output_file_col_num).value = "Term"
    sheet.cell(row=output_file_row_num, column=output_file_col_num).alignment = Alignment(
        horizontal='left', vertical='top')
    sheet.cell(row=output_file_row_num, column=output_file_col_num).border = Border(
        top=Side(style=cell_border), bottom=Side(style=cell_border))
    sheet.column_dimensions[chr(
        col_num_in_alphabets+output_file_col_num)].width = float(7.73)
    output_file_col_num += 1
    # Limit
    sheet.cell(row=output_file_row_num, column=output_file_col_num).font = Font(
        name=text_style, size=string_size, bold=True)
    sheet.cell(row=output_file_row_num,
               column=output_file_col_num).value = "Limit"
    sheet.cell(row=output_file_row_num, column=output_file_col_num).alignment = Alignment(
        horizontal='right', vertical='top')
    sheet.cell(row=output_file_row_num, column=output_file_col_num).border = Border(
        top=Side(style=cell_border), bottom=Side(style=cell_border))
    sheet.column_dimensions[chr(
        col_num_in_alphabets+output_file_col_num)].width = float(12.91)

    num_of_balance_weeks = 5
    # Month Overdue
    for i in range(1, num_of_balance_weeks):

        sheet.cell(row=output_file_row_num, column=output_file_col_num +
                   i).font = Font(name=text_style, size=num_size, bold=True)

        sheet.cell(row=output_file_row_num, column=output_file_col_num +
                   i).border = Border(top=Side(style=cell_border), bottom=Side(style=cell_border))
        sheet.cell(row=output_file_row_num, column=output_file_col_num +
                   i).alignment = Alignment(horizontal='right', vertical='top', wrap_text=False)
        sheet.cell(row=output_file_row_num, column=output_file_col_num +
                   i).value = today.strftime('%b')
        sheet.column_dimensions[chr(
            col_num_in_alphabets+output_file_col_num+i)].width = float(12.91)
        today = today.replace(day=1) - datetime.timedelta(days=1)
        i += 1

    output_file_col_num += num_of_balance_weeks
    # Older
    sheet.cell(row=output_file_row_num, column=output_file_col_num).font = Font(
        name=text_style, size=num_size, bold=True)
    sheet.cell(row=output_file_row_num,
               column=output_file_col_num).value = "Older"
    sheet.cell(row=output_file_row_num, column=output_file_col_num).alignment = Alignment(
        horizontal='right', vertical='top')
    sheet.cell(row=output_file_row_num, column=output_file_col_num).border = Border(
        top=Side(style=cell_border), bottom=Side(style=cell_border))
    sheet.column_dimensions[chr(
        col_num_in_alphabets+output_file_col_num)].width = float(12.91)
    output_file_col_num += 1
    # Total
    sheet.cell(row=output_file_row_num, column=output_file_col_num).font = Font(
        name=text_style, size=num_size, bold=True)
    sheet.cell(row=output_file_row_num,
               column=output_file_col_num).value = "Total"
    sheet.cell(row=output_file_row_num, column=output_file_col_num).alignment = Alignment(
        horizontal='right', vertical='top')
    sheet.cell(row=output_file_row_num, column=output_file_col_num).border = Border(
        top=Side(style=cell_border), bottom=Side(style=cell_border))
    sheet.column_dimensions[chr(
        col_num_in_alphabets+output_file_col_num)].width = float(12.91)

    # write aging report content
    Report.generate_monthly_aging_report(report_file_name, wb, sheet, CUSTOMERS,
                                         customerJsonPath, AGED_RECEIVABLES_SUMMARY, output_file_row_num, text_style, 11, 45)
    os.remove(customerJsonPath)


if __name__ == "__main__":
    main()
