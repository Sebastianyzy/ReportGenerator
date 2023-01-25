import csv
import json
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font
import Report


def get_filename():
    return 


def main():
    CUSTOMERS = r'Customers.csv'
    AGED_RECEIVABLES_SUMMARY = r"Planway_Poultry_Inc__-_Aged_Receivables_Summary.xlsx"
    customerJsonPath = r'Customers.json'
    report_file_name = "PlanwayIndividualAgeingReport.xlsx"

    # create a workbook to store aging report 
    wb = Workbook()
    sheet = wb.active
    output_file_row_num = 1

    # write title 
    # font style = "Arial", size = 20, bold
    sheet.cell(row=output_file_row_num, column=1).font = Font(
        name="Arial", size=20, bold=True)
    sheet.cell(row=output_file_row_num,
               column=1).value = "Aged Receivables Summary [INDIVIDUAL]"
    output_file_row_num += 1
    # font style = "Arial", size = 14, no bold
    sheet.cell(row=output_file_row_num, column=1).font = Font(
        name="Arial", size=14, bold=False)
    sheet.cell(row=output_file_row_num,
               column=1).value = "PLANWAY POULTRY INC."
    output_file_row_num += 1
    sheet.cell(row=output_file_row_num, column=1).font = Font(
        name="Arial", size=14, bold=False)
    #Planway aging report is the most recent wednesday
    sheet.cell(row=output_file_row_num,
               column=1).value = "Effective as at EOD {}".format(Report.get_most_recent_weekdays(3))
    output_file_row_num += 1

    sheet.cell(row=output_file_row_num, column=1).font = Font(
        name="Arial", size=14, bold=False)
    sheet.cell(row=output_file_row_num, column=1).value = "Ageing by due date"
    output_file_row_num += 1

    # write a filter row
    # Name
    sheet.cell(row=output_file_row_num, column=2).font = Font(
        name="Arial", size=11, bold=True)
    sheet.cell(row=output_file_row_num, column=2).value = "Name"
    # Code
    sheet.cell(row=output_file_row_num, column=3).font = Font(
        name="Arial", size=11, bold=True)
    sheet.cell(row=output_file_row_num, column=3).value = "Code"
    # Term
    sheet.cell(row=output_file_row_num, column=4).font = Font(
        name="Arial", size=11, bold=True)
    sheet.cell(row=output_file_row_num, column=4).value = "Term"
    # Limit
    sheet.cell(row=output_file_row_num, column=5).font = Font(
        name="Arial", size=11, bold=True)
    sheet.cell(row=output_file_row_num, column=5).value = "Limit"
    #Current
    sheet.cell(row=output_file_row_num, column=6).font = Font(
        name="Arial", size=10, bold=True)
    sheet.cell(row=output_file_row_num, column=6).value = "Current"
    # Week Overdue
    for i in range(1, 5):
        sheet.cell(row=output_file_row_num, column=6+i).font = Font(name="Arial", size=10, bold=True)
        sheet.cell(row=output_file_row_num, column=6+i).value = f"{i} Week\nOverdue" if i == 1 else f"{i} Weeks\nOverdue"
    # Older    
    sheet.cell(row=output_file_row_num, column=11).font = Font(
        name="Arial", size=10, bold=True)
    sheet.cell(row=output_file_row_num, column=11).value = "Older"  
    # Total 
    sheet.cell(row=output_file_row_num, column=12).font = Font(
        name="Arial", size=10, bold=True)
    sheet.cell(row=output_file_row_num, column=12).value = "Total"   

    # write aging report content
    Report.generate_aging_report(report_file_name , wb, sheet, CUSTOMERS,
                           customerJsonPath, AGED_RECEIVABLES_SUMMARY, output_file_row_num, "Arial", 11)
    os.remove(customerJsonPath)


if __name__ == "__main__":
    main()
