import csv
import json
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from datetime import date
from datetime import timedelta

# This function gets the most recent week days, input integer 1...7, 1 = Monday...7 = Sunday


def get_most_recent_weekdays(day):
    today = date.today()
    offset = (today.weekday() - (day-1))
    last_weekday = today - timedelta(days=offset)
    return "{} {} {}".format(last_weekday.strftime('%a'), last_weekday.strftime("%b"), last_weekday.strftime("%d"))

# Return the file name which contain keyword


def get_filename(keyword):
    current_dir = os.listdir()
    for i in current_dir:
        if keyword.lower() in i.lower():
            return str(i)
    return ""


# Function to convert a CSV to JSON
# Takes the file paths as arguments
def make_json(csvFilePath, jsonFilePath, primaryKey):
    # create a dictionary
    data = {}
    # Open a csv reader called DictReader
    with open(csvFilePath, encoding='utf-8') as csvf:
        csvReader = csv.DictReader(csvf)
        # Convert each row into a dictionary
        # and add it to data
        for rows in csvReader:
            # Assuming a column named 'No' to
            # be the primary key
            key = rows[str(primaryKey)]
            data[key] = rows
    # open a json writer, and use the json.dumps()
    # function to dump data
    with open(jsonFilePath, 'w', encoding='utf-8') as jsonf:
        jsonf.write(json.dumps(data, indent=4))

# Return the attributes of customer object


def load_customer_obj(customer_name, customer_data):
    customer_code = customer_data[customer_name]["\ufeff*Customer Code"]
    customer_type = customer_data[customer_name]["Customer Type"]
    payment_term = customer_data[customer_name]["Payment Terms"]
    credit_limit = customer_data[customer_name]["Credit Limit"]
    salesperson = customer_data[customer_name]["Salesperson"]
    salesperson_prefix = "{}:".format(str(salesperson.split(":", 1)[0]))
    isObsoleted = customer_data[customer_name]["IsObsoleted"]
    return customer_code, customer_type, payment_term, credit_limit, salesperson, salesperson_prefix, isObsoleted

# Construct customer object using the json file


def construct_customer_dictionary(customers, customerJsonPath):
    # call the make_json function
    make_json(customers, customerJsonPath, "*Customer Name")
    with open(customerJsonPath, "r") as content:
        customer_data = json.loads(content.read())
    return customer_data


def generate_aging_report_ind(filename, workbook, sheet, customercsv, customerjson, aged_receivables_summary, line_num, font_style, font_size, row_height):
    customer_data = construct_customer_dictionary(customercsv, customerjson)
    df = pd.DataFrame(pd.read_excel(aged_receivables_summary))
    # starting line
    output_file_row_num = line_num
    num_of_col = len(list(df))
    num_of_row = len(df)

    # the first column of the Aged Receivables Summary, which contain all the customer names
    contact = list(df)[0]
    # the current balance owed by customer
    current = list(df)[1]
    # the balance owed by customer less than a week
    leone_week = list(df)[2]
    # second last column of the file, the balance owed by customer more than the setting weeks
    older = list(df)[num_of_col-2]
    # total balance owed, the last column of the file
    total = list(df)[num_of_col-1]
    cell_font = Font(name=font_style, size=font_size, bold=False)
    cell_border = Border(bottom=Side(style='hair'))  # dotted, hair
    cell_alignment_str = Alignment(horizontal='left',vertical='top',wrap_text=True)
    cell_alignment_num = Alignment(horizontal='right',vertical='top',wrap_text=False)
    height = row_height
    number_format = "0.00"
    # loop through the contact list, and start generate report
    for i in range(0, num_of_row):
        customer_name = str(df[contact][i])
        # check if the customer is in customer database
        if (customer_name in customer_data):
            output_file_row_num += 1
            sheet.row_dimensions[output_file_row_num].height = float(height)

            # get the set of customer attributes from customer data
            customer_code, customer_type, payment_term, credit_limit, salesperson, salesperson_prefix, isObsoleted = load_customer_obj(
                customer_name, customer_data)
            # input salesperson
            c1 = sheet.cell(row=output_file_row_num, column=1)
            c1.font = cell_font
            c1.number_format = number_format
            c1.border = cell_border
            c1.alignment = cell_alignment_str
            c1.value = salesperson_prefix

            # input customer name
            c2 = sheet.cell(row=output_file_row_num, column=2)
            c2.font = cell_font
            c2.number_format = number_format
            c2.border = cell_border
            c2.alignment = cell_alignment_str
            c2.value = customer_name

            # input customer code
            c3 = sheet.cell(row=output_file_row_num, column=3)
            c3.font = cell_font
            c3.number_format = number_format
            c3.border = cell_border
            c3.alignment = cell_alignment_str
            c3.value = customer_code

            # input payment term
            c4 = sheet.cell(row=output_file_row_num, column=4)
            c4.font = cell_font
            c4.number_format = number_format
            c4.border = cell_border
            c4.alignment = cell_alignment_str
            c4.value = payment_term

            # input credit limit
            c5 = sheet.cell(row=output_file_row_num, column=5)
            c5.font = cell_font
            c5.number_format = number_format
            c5.border = cell_border
            c5.alignment = cell_alignment_num
            c5.value = credit_limit

            # input current owed, current owed = current + (<1 week)
            current_owed = df[current][i]
            leone_week_owed = df[leone_week][i]
            c6 = sheet.cell(row=output_file_row_num, column=6)
            c6.font = cell_font
            c6.number_format = number_format
            num = float(current_owed)+float(leone_week_owed)
            c6.border = cell_border
            c6.alignment = cell_alignment_num
            c6.value = '({0:.2f})'.format(
                abs(num)) if num < 0 else '{0:.2f}'.format(num)

            # input the rest of the balance
            k = 7
            for j in range(3, num_of_col-2):
                c = sheet.cell(row=output_file_row_num, column=k)
                c.font = cell_font
                num = float(df[list(df)[j]][i])
                c.border = cell_border
                c.alignment = cell_alignment_num
                c.value = '({0:.2f})'.format(
                    abs(num)) if num < 0 else '{0:.2f}'.format(num)
                k += 1

            # older balance owed
            ck = sheet.cell(row=output_file_row_num, column=k)
            ck.font = cell_font
            ck.number_format = number_format
            num = float(df[older][i])
            ck.border = cell_border
            ck.alignment = cell_alignment_num
            ck.value = '({0:.2f})'.format(
                abs(num)) if num < 0 else '{0:.2f}'.format(num)

            # total balance owed
            ck1 = sheet.cell(row=output_file_row_num, column=k+1)
            ck1.font = cell_font
            ck1.number_format = number_format
            num = float(df[total][i])
            ck1.border = cell_border
            ck1.alignment = cell_alignment_num
            ck1.value = '({0:.2f})'.format(
                abs(num)) if num < 0 else '{0:.2f}'.format(num)
        i += 1
    workbook.save(filename)



def generate_aging_report_full(filename, workbook, sheet, customercsv, customerjson, aged_receivables_summary, line_num, font_style, font_size, row_height):
    customer_data = construct_customer_dictionary(customercsv, customerjson)
    df = pd.DataFrame(pd.read_excel(aged_receivables_summary))
    # starting line
    output_file_row_num = line_num
    num_of_col = len(list(df))
    num_of_row = len(df)


    # the first column of the Aged Receivables Summary, which contain all the customer names
    contact = list(df)[0]
    # the current balance owed by customer
    current = list(df)[1]
    # the balance owed by customer less than a week
    leone_week = list(df)[2]
    # second last column of the file, the balance owed by customer more than the setting weeks
    older = list(df)[num_of_col-2]
    # total balance owed, the last column of the file
    total = list(df)[num_of_col-1]
    cell_font = Font(name=font_style, size=font_size, bold=False)
    cell_border = Border(bottom=Side(style='hair'))  # dotted, hair
    cell_alignment_str = Alignment(horizontal='left',vertical='top',wrap_text=True)
    cell_alignment_num = Alignment(horizontal='right',vertical='top',wrap_text=False)
    height = row_height
    number_format = "0.00"
    # loop through the contact list, and start generate report
    for i in range(0, num_of_row):
        customer_name = str(df[contact][i])
        # check if the customer is in customer database
        if (customer_name in customer_data):
            output_file_row_num += 1
            sheet.row_dimensions[output_file_row_num].height = float(height)

            # get the set of customer attributes from customer data
            customer_code, customer_type, payment_term, credit_limit, salesperson, salesperson_prefix, isObsoleted = load_customer_obj(
                customer_name, customer_data)
            # input code
            c1 = sheet.cell(row=output_file_row_num, column=1)
            c1.font = cell_font
            c1.number_format = number_format
            c1.border = cell_border
            c1.alignment = cell_alignment_str
            c1.value = customer_code

            # input customer name
            c2 = sheet.cell(row=output_file_row_num, column=2)
            c2.font = cell_font
            c2.number_format = number_format
            c2.border = cell_border
            c2.alignment = cell_alignment_str
            c2.value = customer_name

            # input customer term
            c3 = sheet.cell(row=output_file_row_num, column=3)
            c3.font = cell_font
            c3.number_format = number_format
            c3.border = cell_border
            c3.alignment = cell_alignment_str
            c3.value = payment_term

            # input payment limit
            c4 = sheet.cell(row=output_file_row_num, column=4)
            c4.font = cell_font
            c4.number_format = number_format
            c4.border = cell_border
            c4.alignment = cell_alignment_num
            c4.value = credit_limit

            # # input credit limit
            # c5 = sheet.cell(row=output_file_row_num, column=5)
            # c5.font = cell_font
            # c5.number_format = number_format
            # c5.border = cell_border
            # c5.alignment = cell_alignment_num
            # c5.value = credit_limit

            # input current owed, current owed = current + (<1 week)
            current_owed = df[current][i]
            leone_week_owed = df[leone_week][i]
            c5 = sheet.cell(row=output_file_row_num, column=5)
            c5.font = cell_font
            c5.number_format = number_format
            num = float(current_owed)+float(leone_week_owed)
            c5.border = cell_border
            c5.alignment = cell_alignment_num
            c5.value = '({0:.2f})'.format(
                abs(num)) if num < 0 else '{0:.2f}'.format(num)

            # input the rest of the balance
            k = 6
            for j in range(3, num_of_col-2):
                c = sheet.cell(row=output_file_row_num, column=k)
                c.font = cell_font
                num = float(df[list(df)[j]][i])
                c.border = cell_border
                c.alignment = cell_alignment_num
                c.value = '({0:.2f})'.format(
                    abs(num)) if num < 0 else '{0:.2f}'.format(num)
                k += 1

            # older balance owed
            ck = sheet.cell(row=output_file_row_num, column=k)
            ck.font = cell_font
            ck.number_format = number_format
            num = float(df[older][i])
            ck.border = cell_border
            ck.alignment = cell_alignment_num
            ck.value = '({0:.2f})'.format(
                abs(num)) if num < 0 else '{0:.2f}'.format(num)

            # total balance owed
            ck1 = sheet.cell(row=output_file_row_num, column=k+1)
            ck1.font = cell_font
            ck1.number_format = number_format
            num = float(df[total][i])
            ck1.border = cell_border
            ck1.alignment = cell_alignment_num
            ck1.value = '({0:.2f})'.format(
                abs(num)) if num < 0 else '{0:.2f}'.format(num)
        i += 1
    workbook.save(filename)    
